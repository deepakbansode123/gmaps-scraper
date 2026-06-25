"""
Google Maps Business Scraper — Pro Edition
==========================================
Self-contained Streamlit app — no external helper module required.

Features
--------
* Hybrid scraping: fast feed extraction + optional parallel detail-page fetch
* Worldwide city presets (India / USA / UK / Canada / Australia / UAE / custom)
* Rich fields: name, rating, reviews, category, address, phone, website,
  plus_code, latitude, longitude, today's hours, city, maps_url
* Parallel detail extraction (2-4 browser tabs) for speed
* Anti-bot measures: realistic UA, locale, random delays, retries
* Multi-format export: CSV / Excel (xlsx) / JSON
* Live log, per-city progress, stats dashboard, in-app filtering
* Dedup by name + address; safe cancellation

Run locally:
    streamlit run app.py
"""

from __future__ import annotations

import csv
import io
import json
import queue
import random
import re
import sys
import threading
import time
import subprocess
from pathlib import Path
from datetime import datetime
from urllib.parse import quote_plus

# Install Chromium for Playwright (silently — local installs are skipped)
subprocess.run(
    [sys.executable, "-m", "playwright", "install", "chromium"],
    capture_output=True,
    check=False,
)

import streamlit as st
from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout

# ─────────────────────────────────────────────────────────────────────────────
# Page config
# ─────────────────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Google Maps Scraper — Pro",
    page_icon="🗺️",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─────────────────────────────────────────────────────────────────────────────
# Selectors (Google Maps DOM as of 2024-2025)
# ─────────────────────────────────────────────────────────────────────────────
SEL = {
    # Feed / cards
    "feed":          'div[role="feed"]',
    "card_link":     'a.hfpxzc',                          # listing card anchor
    "card_name":     'div.qBF1Pd, .fontHeadlineSmall',
    "card_rating":   'span.MW4etd',
    "card_reviews":  'span.UY7F9',
    "card_category": 'span.YhemCb',
    "card_address":  'button[data-item-id="address"] .rFm3Rc, .W4Efnf',
    # Detail page
    "name":          'h1.DUwDvf',
    "rating":        'div.F7nice span[aria-hidden="true"]',
    "reviews":       'div.F7nice button[aria-label]',
    "category":      'button[jsaction*="category"] span, button[jsaction*="pane.rating.category"]',
    "address":       'button[data-item-id="address"] div.Io6YIf, [data-item-id="address"] .Io6YIf',
    "phone":         'button[data-item-id^="phone:"] div.Io6YIf, [data-item-id^="phone:"] .Io6YIf',
    "website":       'a[data-item-id="authority"]',
    "plus_code":     'button[data-item-id="oloc"] div.Io6YIf, [data-item-id="oloc"] .Io6YIf',
    "hours":         'div.t39EBfGU tbody tr:first-child td:last-child, div.t39EBfGU span',
}

# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────
USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36",
]

COORD_RE = re.compile(r"!3d(-?\d+\.\d+)!4d(-?\d+\.\d+)")


def text_or_blank(page, selector: str, timeout: int = 1500) -> str:
    """Try to read text from a selector; return '' if not present."""
    try:
        el = page.query_selector(selector)
        if not el:
            return ""
        txt = (el.inner_text() or "").strip()
        return re.sub(r"\s+", " ", txt)
    except Exception:
        return ""


def attr_or_blank(page, selector: str, attr: str, timeout: int = 1500) -> str:
    try:
        el = page.query_selector(selector)
        if not el:
            return ""
        return (el.get_attribute(attr) or "").strip()
    except Exception:
        return ""


def parse_coords_from_url(url: str) -> tuple[str, str]:
    m = COORD_RE.search(url or "")
    if m:
        return m.group(1), m.group(2)
    return "", ""


def parse_reviews_count(text: str) -> str:
    """Extract clean review count from text like '(1,234)' or '1.2K reviews'."""
    if not text:
        return ""
    m = re.search(r"([\d.,]+\s*[KkMm]?)", text)
    return m.group(1).replace("(", "").replace(")", "").strip() if m else ""


def scroll_feed(page, target_count: int):
    """Scroll the left feed until target_count cards are loaded or feed ends."""
    last_count = 0
    same_rounds = 0
    for _ in range(60):  # safety cap
        cards = page.query_selector_all(SEL["card_link"])
        if len(cards) >= target_count:
            break
        if len(cards) == last_count:
            same_rounds += 1
            if same_rounds >= 4:
                break  # no new content for 4 rounds → end of feed
        else:
            same_rounds = 0
        last_count = len(cards)

        # scroll the feed container
        try:
            page.evaluate(
                """() => {
                    const f = document.querySelector('div[role="feed"]');
                    if (f) f.scrollBy(0, 1200);
                    else window.scrollBy(0, 1200);
                }"""
            )
        except Exception:
            pass
        time.sleep(random.uniform(0.4, 0.8))


def extract_feed_cards(page) -> list[dict]:
    """Fast extraction from feed cards (no detail-page navigation)."""
    rows = []
    cards = page.query_selector_all(SEL["card_link"])
    for card in cards:
        try:
            href = card.get_attribute("href") or ""
            name = (card.get_attribute("aria-label") or "").strip()
            if not name:
                inner = card.query_selector(SEL["card_name"])
                name = inner.inner_text().strip() if inner else ""
            # parent container often holds rating/category
            parent = card.evaluate_handle('el => el.closest("div[role=article]") || el.parentElement.parentElement')
            rating = ""
            reviews = ""
            category = ""
            address = ""
            try:
                rt = parent.query_selector(SEL["card_rating"])
                rating = rt.inner_text().strip() if rt else ""
            except Exception:
                pass
            try:
                rv = parent.query_selector(SEL["card_reviews"])
                reviews = parse_reviews_count(rv.inner_text() if rv else "")
            except Exception:
                pass
            try:
                cat = parent.query_selector(SEL["card_category"])
                category = cat.inner_text().strip() if cat else ""
            except Exception:
                pass
            try:
                ad = parent.query_selector(SEL["card_address"])
                address = ad.inner_text().strip() if ad else ""
            except Exception:
                pass

            lat, lng = parse_coords_from_url(href)
            rows.append({
                "name": name,
                "rating": rating,
                "reviews": reviews,
                "category": category,
                "address": address,
                "phone": "",
                "website": "",
                "plus_code": "",
                "latitude": lat,
                "longitude": lng,
                "hours_today": "",
                "maps_url": href,
            })
        except Exception:
            continue
    return rows


def enrich_from_detail(page, row: dict) -> dict:
    """Visit a detail page and fill in missing fields (phone, website, etc.)."""
    try:
        page.goto(row["maps_url"], wait_until="domcontentloaded", timeout=30000)
        try:
            page.wait_for_selector(SEL["name"], timeout=10000)
        except PWTimeout:
            return row
        time.sleep(random.uniform(0.4, 0.9))

        # Refresh name (more accurate from detail page)
        dname = text_or_blank(page, SEL["name"])
        if dname:
            row["name"] = dname
        if not row["rating"]:
            row["rating"] = text_or_blank(page, SEL["rating"])
        if not row["reviews"]:
            rv_text = ""
            try:
                rv_el = page.query_selector(SEL["reviews"])
                if rv_el:
                    rv_text = rv_el.get_attribute("aria-label") or rv_el.inner_text() or ""
            except Exception:
                pass
            row["reviews"] = parse_reviews_count(rv_text)
        if not row["category"]:
            row["category"] = text_or_blank(page, SEL["category"])
        if not row["address"]:
            row["address"] = text_or_blank(page, SEL["address"])
        row["phone"] = text_or_blank(page, SEL["phone"])
        row["website"] = attr_or_blank(page, SEL["website"], "href")
        if not row["plus_code"]:
            row["plus_code"] = text_or_blank(page, SEL["plus_code"])
        if not row["hours_today"]:
            row["hours_today"] = text_or_blank(page, SEL["hours"])

        # coords from URL (detail page URL often includes them)
        lat, lng = parse_coords_from_url(page.url)
        if lat and lng:
            row["latitude"] = lat
            row["longitude"] = lng
        row["maps_url"] = page.url
    except Exception:
        pass
    return row


# ─────────────────────────────────────────────────────────────────────────────
# City presets — worldwide
# ─────────────────────────────────────────────────────────────────────────────
COUNTRY_PRESETS = {
    "India — Maharashtra": ["Mumbai", "Pune", "Nagpur", "Nashik", "Aurangabad", "Kolhapur", "Solapur", "Amravati", "Sangli", "Ahmednagar"],
    "India — Gujarat": ["Ahmedabad", "Surat", "Vadodara", "Rajkot", "Bhavnagar", "Jamnagar", "Junagadh", "Gandhinagar"],
    "India — Rajasthan": ["Jaipur", "Jodhpur", "Udaipur", "Kota", "Bikaner", "Ajmer", "Alwar"],
    "India — Uttar Pradesh": ["Lucknow", "Kanpur", "Varanasi", "Agra", "Meerut", "Prayagraj", "Bareilly", "Moradabad"],
    "India — Bihar": ["Patna", "Gaya", "Muzaffarpur", "Bhagalpur", "Darbhanga", "Motihari"],
    "India — Punjab": ["Ludhiana", "Amritsar", "Jalandhar", "Patiala", "Bathinda", "Mohali"],
    "India — Haryana": ["Ambala", "Hisar", "Rohtak", "Karnal", "Panipat", "Yamunanagar", "Sirsa"],
    "India — Madhya Pradesh": ["Bhopal", "Indore", "Jabalpur", "Gwalior", "Ujjain", "Rewa", "Sagar"],
    "India — Delhi NCR": ["Delhi", "Noida", "Gurgaon", "Ghaziabad", "Faridabad", "Greater Noida"],
    "India — Karnataka": ["Bangalore", "Mysore", "Hubli", "Mangalore", "Belgaum", "Davangere"],
    "India — Tamil Nadu": ["Chennai", "Coimbatore", "Madurai", "Salem", "Tiruchirappalli", "Tirunelveli"],
    "India — Telangana & AP": ["Hyderabad", "Visakhapatnam", "Vijayawada", "Guntur", "Nellore", "Kurnool", "Warangal"],
    "India — Chhattisgarh": ["Raipur", "Bilaspur", "Bhilai", "Korba", "Durg", "Rajnandgaon"],
    "India — Odisha": ["Bhubaneswar", "Cuttack", "Berhampur", "Rourkela", "Sambalpur"],
    "India — West Bengal": ["Kolkata", "Howrah", "Asansol", "Siliguri", "Durgapur", "Kharagpur"],
    "India — Kerala": ["Kochi", "Thiruvananthapuram", "Kozhikode", "Thrissur", "Kollam", "Kannur"],
    "USA — California": ["Los Angeles", "San Francisco", "San Diego", "San Jose", "Sacramento", "Fresno", "Oakland"],
    "USA — Texas": ["Houston", "Dallas", "Austin", "San Antonio", "Fort Worth", "El Paso"],
    "USA — New York": ["New York City", "Buffalo", "Rochester", "Yonkers", "Syracuse", "Albany"],
    "USA — Florida": ["Miami", "Orlando", "Tampa", "Jacksonville", "Fort Lauderdale", "St. Petersburg"],
    "USA — Illinois": ["Chicago", "Aurora", "Naperville", "Springfield", "Peoria"],
    "UK — England": ["London", "Manchester", "Birmingham", "Leeds", "Liverpool", "Bristol", "Sheffield", "Newcastle"],
    "UK — Scotland": ["Edinburgh", "Glasgow", "Aberdeen", "Dundee", "Inverness"],
    "Canada": ["Toronto", "Montreal", "Vancouver", "Calgary", "Ottawa", "Edmonton", "Winnipeg", "Quebec City"],
    "Australia": ["Sydney", "Melbourne", "Brisbane", "Perth", "Adelaide", "Canberra", "Gold Coast", "Newcastle"],
    "UAE": ["Dubai", "Abu Dhabi", "Sharjah", "Ajman", "Al Ain", "Ras Al Khaimah"],
    "Singapore": ["Singapore"],
    "Germany": ["Berlin", "Munich", "Hamburg", "Frankfurt", "Cologne", "Stuttgart", "Düsseldorf"],
    "France": ["Paris", "Marseille", "Lyon", "Toulouse", "Nice", "Nantes", "Bordeaux"],
}

SEARCH_PRESETS = [
    "cattle feed dealers",
    "poultry feed dealers",
    "animal feed suppliers",
    "fertilizer dealers",
    "pesticide dealers",
    "seeds dealers",
    "agro chemicals dealers",
    "dairy equipment suppliers",
    "tractor dealers",
    "veterinary medicine shops",
    "irrigation equipment dealers",
    "farm machinery dealers",
    "cold storage services",
    "food processing units",
    "grain traders",
    "restaurants",
    "hotels",
    "gyms",
    "real estate agents",
    "car dealers",
    "doctors",
    "dentists",
    "schools",
    "coffee shops",
    "auto repair shops",
    "pharmacies",
    "salons",
    "lawyers",
    "plumbers",
    "electricians",
]

# ─────────────────────────────────────────────────────────────────────────────
# Session state
# ─────────────────────────────────────────────────────────────────────────────
_defaults = {
    "running": False,
    "results": [],
    "log": [],
    "scraped_count": 0,
    "total_expected": 0,
    "city_progress": {},  # city -> count
    "q": None,
    "city_list": [],
    "start_time": None,
    "stop_flag": False,
}
for k, v in _defaults.items():
    if k not in st.session_state:
        st.session_state[k] = v


# ─────────────────────────────────────────────────────────────────────────────
# Scraper worker
# ─────────────────────────────────────────────────────────────────────────────
def scrape_worker(cities, term, max_per_city, enrich, parallel_tabs, q, stop_flag):
    """
    Hybrid scraper.
      - Phase 1: load feed, extract basic data (name, rating, reviews,
        category, address, coords) — very fast.
      - Phase 2 (optional): visit detail pages in parallel to fetch
        phone / website / plus_code / hours.
    """
    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(
                headless=True,
                args=[
                    "--no-sandbox",
                    "--disable-blink-features=AutomationControlled",
                    "--disable-dev-shm-usage",
                    "--disable-gpu",
                ],
            )
            context = browser.new_context(
                locale="en-US",
                viewport={"width": 1366, "height": 900},
                user_agent=random.choice(USER_AGENTS),
                timezone_id="America/New_York",
            )
            # tweak navigator.webdriver
            context.add_init_script(
                "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"
            )

            main_page = context.new_page()

            all_rows: list[dict] = []

            for city in cities:
                if stop_flag["stop"]:
                    q.put(("log", "⏹  Stopped by user."))
                    break

                q.put(("log", f"\n🔍  {city}  —  searching…"))
                q.put(("city_start", city))

                url = (
                    "https://www.google.com/maps/search/"
                    + quote_plus(f"{term} in {city}")
                )

                try:
                    main_page.goto(url, wait_until="domcontentloaded", timeout=60000)

                    # accept cookies if present (EU)
                    try:
                        main_page.click(
                            'button[aria-label*="Accept"], button:has-text("Accept all"), button:has-text("Reject all")',
                            timeout=2500,
                        )
                    except PWTimeout:
                        pass

                    # wait for feed
                    try:
                        main_page.wait_for_selector(SEL["feed"], timeout=20000)
                    except PWTimeout:
                        q.put(("log", f"   ⚠️  {city}: no results (CAPTCHA / blocked?). Skipping."))
                        q.put(("city_done", (city, 0)))
                        continue

                    # Phase 1 — scroll + extract from feed
                    scroll_feed(main_page, max_per_city)
                    rows = extract_feed_cards(main_page)[:max_per_city]

                    q.put(("log", f"   📋  {city}: {len(rows)} listings found in feed."))

                    # Phase 2 — optional detail enrichment (parallel)
                    if enrich and rows:
                        pages_pool = [main_page]
                        for _ in range(max(0, parallel_tabs - 1)):
                            pages_pool.append(context.new_page())

                        idx = 0
                        completed = 0
                        lock = threading.Lock()

                        def worker(page_obj):
                            nonlocal idx, completed
                            while True:
                                if stop_flag["stop"]:
                                    return
                                with lock:
                                    if idx >= len(rows):
                                        return
                                    cur = rows[idx]
                                    idx += 1
                                enrich_from_detail(page_obj, cur)
                                with lock:
                                    completed += 1
                                q.put(("tick", 1))
                                q.put(("city_tick", city))
                                time.sleep(random.uniform(0.2, 0.5))

                        threads = [threading.Thread(target=worker, args=(pg,), daemon=True) for pg in pages_pool]
                        for t in threads:
                            t.start()
                        for t in threads:
                            t.join()

                        for extra in pages_pool[1:]:
                            try:
                                extra.close()
                            except Exception:
                                pass
                        rows = [r for r in rows if r.get("name")]
                        q.put(("log", f"   ✅  {city}: enriched {completed} listings."))
                    else:
                        # just count ticks
                        for _ in rows:
                            q.put(("tick", 1))
                            q.put(("city_tick", city))
                        q.put(("log", f"   ✅  {city}: {len(rows)} listings (feed-only mode)."))

                    all_rows.extend(rows)
                    q.put(("city_done", (city, len(rows))))

                    # small inter-city delay to be polite
                    time.sleep(random.uniform(1.0, 2.0))

                except Exception as e:
                    q.put(("log", f"   ❌  {city}: {e}"))
                    q.put(("city_done", (city, 0)))

            try:
                browser.close()
            except Exception:
                pass

        # Dedup by (name lower + address lower)
        seen_k, uniq = set(), []
        for r in all_rows:
            key = (
                (r.get("name") or "").lower().strip(),
                (r.get("address") or "").lower().strip()[:60],
            )
            if r.get("name") and key not in seen_k:
                seen_k.add(key)
                uniq.append(r)

        q.put(("done", uniq))

    except Exception as e:
        q.put(("error", str(e)))


# ─────────────────────────────────────────────────────────────────────────────
# Export helpers
# ─────────────────────────────────────────────────────────────────────────────
EXPORT_FIELDS = [
    "name", "rating", "reviews", "category", "address",
    "phone", "website", "plus_code", "latitude", "longitude",
    "hours_today", "city", "maps_url",
]


def make_csv(rows, city_filter=None):
    show = rows if city_filter in (None, "All") else [r for r in rows if r.get("city") == city_filter]
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=EXPORT_FIELDS, extrasaction="ignore")
    writer.writeheader()
    writer.writerows(show)
    return buf.getvalue().encode("utf-8-sig")


def make_json(rows, city_filter=None):
    show = rows if city_filter in (None, "All") else [r for r in rows if r.get("city") == city_filter]
    return json.dumps(show, indent=2, ensure_ascii=False).encode("utf-8")


def make_xlsx(rows, city_filter=None):
    """Build an .xlsx file using openpyxl (lazy import)."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError as e:
        raise RuntimeError("openpyxl not installed. Add it to requirements.txt") from e

    show = rows if city_filter in (None, "All") else [r for r in rows if r.get("city") == city_filter]
    wb = Workbook()
    ws = wb.active
    ws.title = "Listings"
    headers = EXPORT_FIELDS
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="1A73E8")
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for r in show:
        ws.append([r.get(k, "") for k in EXPORT_FIELDS])
    # auto column width
    for col_idx, h in enumerate(headers, start=1):
        max_len = len(str(h))
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx, values_only=True):
            for v in row:
                if v is not None:
                    max_len = max(max_len, min(60, len(str(v))))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max_len + 2
    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def safe_filename(s: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", s.lower()).strip("_") or "gmaps_export"


# ─────────────────────────────────────────────────────────────────────────────
# Sidebar — configuration
# ─────────────────────────────────────────────────────────────────────────────
with st.sidebar:
    st.title("🗺️ GMaps Scraper Pro")
    st.caption("Google Maps business data extractor")
    st.divider()

    # 1. Search Term
    st.subheader("1️⃣  Search Term")
    preset = st.selectbox("Quick presets", ["— custom —"] + SEARCH_PRESETS)
    default_term = "" if preset == "— custom —" else preset
    term = st.text_input(
        "Search term",
        value=default_term,
        placeholder="e.g. cattle feed dealers",
        label_visibility="collapsed",
    )
    st.divider()

    # 2. Cities
    st.subheader("2️⃣  Cities")
    country_pick = st.selectbox("Region preset", ["— none —"] + list(COUNTRY_PRESETS.keys()))
    raw_cities = st.text_area(
        "City list (one per line)",
        value="\n".join(st.session_state.city_list),
        height=150,
        placeholder="Mumbai\nPune\nNagpur\n…",
        label_visibility="collapsed",
    )
    cities = [c.strip() for c in raw_cities.splitlines() if c.strip()]

    c1, c2 = st.columns(2)
    if c1.button("➕ Add region", use_container_width=True):
        if country_pick != "— none —":
            merged = list(cities)
            for c in COUNTRY_PRESETS[country_pick]:
                if c not in merged:
                    merged.append(c)
            st.session_state.city_list = merged
            st.rerun()
    if c2.button("🗑️  Clear", use_container_width=True):
        st.session_state.city_list = []
        st.rerun()

    st.caption(f"**{len(cities)}** cities selected")
    st.divider()

    # 3. Settings
    st.subheader("3️⃣  Settings")
    max_per = st.slider("Max results per city", 10, 200, 60, step=10)
    enrich = st.toggle(
        "🔍  Enrich via detail pages",
        value=True,
        help="Visits each listing page to fetch phone, website, plus_code, hours. Slower but richer.",
    )
    parallel_tabs = st.slider(
        "Parallel tabs (for enrichment)", 1, 4, 2,
        disabled=not enrich,
        help="More tabs = faster but higher risk of being rate-limited.",
    )
    st.divider()

    ready = bool(term.strip() and cities)

    if not st.session_state.running:
        if st.button("🚀  Start Scraping", type="primary",
                     use_container_width=True, disabled=not ready):
            q = queue.Queue()
            st.session_state.q = q
            st.session_state.running = True
            # IMPORTANT: store the stop dict in session_state so the Stop
            # button can mutate the SAME object the worker is polling.
            st.session_state.stop_flag_obj = {"stop": False}
            st.session_state.results = []
            st.session_state.scraped_count = 0
            st.session_state.city_progress = {c: 0 for c in cities}
            st.session_state.total_expected = len(cities) * max_per
            st.session_state.start_time = time.time()
            st.session_state.log = [
                f"Started  : {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
                f"Search   : {term}",
                f"Cities   : {', '.join(cities)}",
                f"Max/city : {max_per}",
                f"Enrich   : {enrich}  |  Tabs: {parallel_tabs}",
                "─" * 50,
            ]
            threading.Thread(
                target=scrape_worker,
                args=(cities, term.strip(), max_per, enrich, parallel_tabs, q,
                      st.session_state.stop_flag_obj),
                daemon=True,
            ).start()
            st.rerun()

        if not ready:
            if not term.strip():
                st.warning("Enter a search term.")
            if not cities:
                st.warning("Add at least 1 city.")
    else:
        sc1, sc2 = st.columns(2)
        if sc1.button("⏹  Stop", use_container_width=True, type="secondary"):
            # Mutate the SAME dict the worker is polling
            if st.session_state.get("stop_flag_obj"):
                st.session_state.stop_flag_obj["stop"] = True
            st.session_state.log.append("⏹  Stop requested — finishing current items…")
            st.rerun()
        sc2.button("⏳  Running…", disabled=True, use_container_width=True)


# ─────────────────────────────────────────────────────────────────────────────
# Main area
# ─────────────────────────────────────────────────────────────────────────────
st.header("Google Maps Business Scraper", divider="gray")
st.caption("Configure from the sidebar → press **Start Scraping**.")

# Poll worker queue
needs_rerun = False
if st.session_state.running and st.session_state.q:
    q = st.session_state.q
    while not q.empty():
        try:
            msg, payload = q.get_nowait()
        except queue.Empty:
            break
        needs_rerun = True
        if msg == "log":
            st.session_state.log.append(payload)
        elif msg == "tick":
            st.session_state.scraped_count += payload
        elif msg == "city_tick":
            st.session_state.city_progress[payload] = st.session_state.city_progress.get(payload, 0) + 1
        elif msg == "city_start":
            st.session_state.city_progress[payload] = 0
        elif msg == "city_done":
            city_name, cnt = payload
            st.session_state.city_progress[city_name] = cnt
        elif msg == "done":
            st.session_state.running = False
            st.session_state.results = payload
            elapsed = time.time() - (st.session_state.start_time or time.time())
            st.session_state.log += [
                "─" * 50,
                f"✅  Done!  {len(payload)} unique listings in {elapsed:.1f}s",
                f"📞  Phone  : {sum(1 for r in payload if r.get('phone'))}",
                f"🌐  Website: {sum(1 for r in payload if r.get('website'))}",
                f"⭐  Rating : {sum(1 for r in payload if r.get('rating'))}",
                f"📍  Coords : {sum(1 for r in payload if r.get('latitude'))}",
            ]
        elif msg == "error":
            st.session_state.running = False
            st.session_state.log.append(f"\n❌  Error: {payload}")

# Progress
if st.session_state.total_expected > 0:
    pct = min(st.session_state.scraped_count / max(1, st.session_state.total_expected), 1.0)
    if st.session_state.running:
        elapsed = time.time() - (st.session_state.start_time or time.time())
        rate = st.session_state.scraped_count / elapsed if elapsed > 0 else 0
        label = f"Scraping… {st.session_state.scraped_count} / ~{st.session_state.total_expected}  ({rate:.1f}/s)"
    else:
        label = f"Done — {st.session_state.scraped_count} records scraped"
    st.progress(pct, text=label)

# Per-city mini stats
if st.session_state.city_progress and (st.session_state.running or st.session_state.results):
    with st.expander("🏙️  Per-city progress", expanded=st.session_state.running):
        cols = st.columns(min(6, max(1, len(st.session_state.city_progress))))
        for i, (city, cnt) in enumerate(st.session_state.city_progress.items()):
            with cols[i % len(cols)]:
                st.metric(city, cnt)

# Tabs
tab_log, tab_results, tab_export = st.tabs(["📋  Live Log", "📊  Results", "💾  Export"])

with tab_log:
    if st.session_state.log:
        st.code("\n".join(st.session_state.log), language=None)
    else:
        st.info("👈  Configure from the sidebar, then press **Start Scraping**.")

with tab_results:
    results = st.session_state.results
    if results:
        m1, m2, m3, m4, m5 = st.columns(5)
        m1.metric("Total Listings", len(results))
        m2.metric("With Phone", sum(1 for r in results if r.get("phone")))
        m3.metric("With Website", sum(1 for r in results if r.get("website")))
        m4.metric("With Rating", sum(1 for r in results if r.get("rating")))
        m5.metric("With Coords", sum(1 for r in results if r.get("latitude")))

        st.divider()

        # Filters
        all_cities_in_results = sorted({r.get("city", "") for r in results if r.get("city")})
        fcol1, fcol2, fcol3 = st.columns([2, 2, 3])
        with fcol1:
            city_f = st.selectbox("Filter by city", ["All"] + all_cities_in_results, key="res_city_f")
        with fcol2:
            min_rating = st.number_input("Min rating", 0.0, 5.0, 0.0, 0.1, key="res_min_r")
        with fcol3:
            keyword = st.text_input("Keyword in name/category/address", "", key="res_kw")

        show = results
        if city_f != "All":
            show = [r for r in show if r.get("city") == city_f]
        if min_rating > 0:
            show = [
                r for r in show
                if r.get("rating") and re.match(r"^\d+(\.\d+)?$", str(r["rating"]))
                and float(r["rating"]) >= min_rating
            ]
        if keyword.strip():
            kw = keyword.strip().lower()
            show = [
                r for r in show
                if kw in (r.get("name", "") + r.get("category", "") + r.get("address", "")).lower()
            ]

        st.caption(f"Showing **{len(show)}** of {len(results)} listings")

        st.dataframe(
            show,
            use_container_width=True,
            hide_index=True,
            column_config={
                "maps_url":   st.column_config.LinkColumn("Maps",    display_text="🔗 Open"),
                "website":    st.column_config.LinkColumn("Website", display_text="🌐 Open"),
                "rating":     st.column_config.TextColumn("⭐ Rating"),
                "reviews":    st.column_config.NumberColumn("Reviews", format="%d"),
                "latitude":   st.column_config.NumberColumn("Lat", format="%.6f"),
                "longitude":  st.column_config.NumberColumn("Lng", format="%.6f"),
                "phone":      st.column_config.TextColumn("📞 Phone"),
                "plus_code":  st.column_config.TextColumn("📍 Plus Code"),
                "hours_today": st.column_config.TextColumn("🕘 Hours Today"),
            },
            column_order=EXPORT_FIELDS,
        )
    else:
        st.info("Results will appear here after scraping completes.")

with tab_export:
    results = st.session_state.results
    if results:
        all_cities_in_results = sorted({r.get("city", "") for r in results if r.get("city")})
        export_city = st.selectbox(
            "Export filter (city)", ["All"] + all_cities_in_results, key="export_city_f"
        )
        base_name = safe_filename(term)
        if export_city != "All":
            base_name += f"_{safe_filename(export_city)}"

        d1, d2, d3 = st.columns(3)
        d1.download_button(
            "⬇️  Download CSV", data=make_csv(results, export_city),
            file_name=f"{base_name}.csv", mime="text/csv", use_container_width=True,
        )
        d2.download_button(
            "⬇️  Download Excel", data=make_xlsx(results, export_city),
            file_name=f"{base_name}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
        d3.download_button(
            "⬇️  Download JSON", data=make_json(results, export_city),
            file_name=f"{base_name}.json", mime="application/json", use_container_width=True,
        )
        st.caption("Excel export includes a styled header row and frozen top row.")
    else:
        st.info("Run a scrape first — exports will appear here.")

# Auto-refresh while running
if st.session_state.running:
    time.sleep(1.2)
    st.rerun()
elif needs_rerun:
    st.rerun()
